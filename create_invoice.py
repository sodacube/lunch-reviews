import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tax Invoice"

# Define styles
header_font = Font(bold=True)
title_font = Font(bold=True, size=14)
orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
light_orange_fill = PatternFill(start_color="FFE4C4", end_color="FFE4C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
gray_font = Font(color="808080", italic=True)

# Set column widths
col_widths = {'A': 18, 'B': 15, 'C': 35, 'D': 8, 'E': 12, 'F': 10, 'G': 12, 'H': 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "TAX INVOICE"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

# Supplier info section (left side)
supplier_data = [
    ('Supplier Name:', 'Donna Capurso and Dan Cormick'),
    ('Supplier Address:', '9 Kinnaird Crescent, Highbury SA 5089'),
    ('Supplier Email:', 'donnacapurso@gmail.com'),
    ('Supplier Phone:', '410319158'),
    ('ALS Order #:', '60395'),
]

for i, (label, value) in enumerate(supplier_data, start=3):
    ws[f'A{i}'] = label
    ws[f'A{i}'].font = header_font
    ws[f'A{i}'].alignment = Alignment(horizontal='right')
    ws[f'A{i}'].border = thin_border
    ws.merge_cells(f'B{i}:D{i}')
    ws[f'B{i}'] = value
    ws[f'B{i}'].border = thin_border
    for col in ['C', 'D']:
        ws[f'{col}{i}'].border = thin_border

# Invoice info section (right side) - orange background
invoice_data = [
    ('INVOICE NUMBER:', 'DDD003'),
    ('INVOICE DATE:', '24.3.26'),
    ('ABN:', '94676613498'),  # ABN as text - this is typically an 11-digit number
    ('IS GST CHARGED?:', 'NO'),
]

for i, (label, value) in enumerate(invoice_data, start=3):
    ws.merge_cells(f'F{i}:G{i}')
    ws[f'F{i}'] = label
    ws[f'F{i}'].font = header_font
    ws[f'F{i}'].fill = orange_fill
    ws[f'F{i}'].border = thin_border
    ws[f'G{i}'].border = thin_border
    ws[f'H{i}'] = value
    ws[f'H{i}'].fill = orange_fill
    ws[f'H{i}'].border = thin_border
    # Format ABN as text explicitly
    if label == 'ABN:':
        ws[f'H{i}'].number_format = '@'  # Text format

# Invoice to / Deliver to section
ws['A9'] = 'Invoice to:'
ws['A9'].font = header_font
ws['A9'].alignment = Alignment(horizontal='right')
ws['A9'].border = thin_border

ws.merge_cells('B9:D9')
ws['B9'] = 'ALS Library Services Pty Ltd'
ws['B9'].border = thin_border
for col in ['C', 'D']:
    ws[f'{col}9'].border = thin_border

ws.merge_cells('B10:D10')
ws['B10'] = '12 - 14 Tooronga Ave, Edwardstown SA 5039'
ws['B10'].border = thin_border
ws['A10'].border = thin_border
for col in ['C', 'D']:
    ws[f'{col}10'].border = thin_border

ws['E9'] = 'Deliver to:'
ws['E9'].font = header_font
ws['E9'].alignment = Alignment(horizontal='right')
ws['E9'].border = thin_border

ws.merge_cells('F9:H9')
ws['F9'] = 'ALS Library Services Pty Ltd'
ws['F9'].border = thin_border
for col in ['G', 'H']:
    ws[f'{col}9'].border = thin_border

ws.merge_cells('F10:H10')
ws['F10'] = '14B Konando Tce, Edwardstown SA 5052'
ws['F10'].border = thin_border
ws['E10'].border = thin_border
for col in ['G', 'H']:
    ws[f'{col}10'].border = thin_border

# Line items header
headers = ['ISBN', 'Title / Author', '', 'Qty', 'Retail Price', 'Discount', 'Unit Cost', 'Line Total']
header_row = 12
ws.merge_cells(f'B{header_row}:C{header_row}')
for i, header in enumerate(headers, start=1):
    col = get_column_letter(i)
    if header:  # Skip empty (merged) cells
        ws[f'{col}{header_row}'] = header
        ws[f'{col}{header_row}'].font = header_font
    ws[f'{col}{header_row}'].border = thin_border
    if i >= 4:  # Qty onwards - orange fill
        ws[f'{col}{header_row}'].fill = light_orange_fill

# Example row (grayed out)
example_row = 13
ws.merge_cells(f'B{example_row}:C{example_row}')
ws[f'A{example_row}'] = '9784567890123'
ws[f'A{example_row}'].font = gray_font
ws[f'A{example_row}'].border = thin_border
ws[f'B{example_row}'] = 'Example title / Jane Smith'
ws[f'B{example_row}'].font = gray_font
ws[f'B{example_row}'].border = thin_border
ws[f'C{example_row}'].border = thin_border
ws[f'D{example_row}'] = 5
ws[f'D{example_row}'].font = gray_font
ws[f'D{example_row}'].border = thin_border
ws[f'D{example_row}'].alignment = Alignment(horizontal='right')
ws[f'E{example_row}'] = 19.99
ws[f'E{example_row}'].font = gray_font
ws[f'E{example_row}'].border = thin_border
ws[f'E{example_row}'].alignment = Alignment(horizontal='right')
ws[f'E{example_row}'].number_format = '0.00'
ws[f'F{example_row}'] = '40%'
ws[f'F{example_row}'].font = gray_font
ws[f'F{example_row}'].border = thin_border
ws[f'F{example_row}'].alignment = Alignment(horizontal='right')
ws[f'G{example_row}'] = 11.99
ws[f'G{example_row}'].font = gray_font
ws[f'G{example_row}'].border = thin_border
ws[f'G{example_row}'].alignment = Alignment(horizontal='right')
ws[f'G{example_row}'].number_format = '0.00'
ws[f'H{example_row}'] = 59.97
ws[f'H{example_row}'].font = gray_font
ws[f'H{example_row}'].border = thin_border
ws[f'H{example_row}'].alignment = Alignment(horizontal='right')
ws[f'H{example_row}'].number_format = '0.00'

# Data row
data_row = 14
ws.merge_cells(f'B{data_row}:C{data_row}')
ws[f'A{data_row}'] = '9780645649727'
ws[f'A{data_row}'].border = thin_border
ws[f'B{data_row}'] = "Don't Do It Donna! Brushing Teeth, Donna Capurso"
ws[f'B{data_row}'].border = thin_border
ws[f'C{data_row}'].border = thin_border
ws[f'D{data_row}'] = 8
ws[f'D{data_row}'].border = thin_border
ws[f'D{data_row}'].alignment = Alignment(horizontal='right')
ws[f'E{data_row}'] = 8
ws[f'E{data_row}'].border = thin_border
ws[f'E{data_row}'].alignment = Alignment(horizontal='right')
ws[f'E{data_row}'].number_format = '0.00'
ws[f'F{data_row}'] = '10%'
ws[f'F{data_row}'].border = thin_border
ws[f'F{data_row}'].alignment = Alignment(horizontal='right')
ws[f'G{data_row}'] = 7.20
ws[f'G{data_row}'].border = thin_border
ws[f'G{data_row}'].alignment = Alignment(horizontal='right')
ws[f'G{data_row}'].number_format = '0.00'
ws[f'H{data_row}'] = 57.60
ws[f'H{data_row}'].border = thin_border
ws[f'H{data_row}'].alignment = Alignment(horizontal='right')
ws[f'H{data_row}'].number_format = '0.00'

# Empty rows with 0.00 values
for row in range(15, 23):
    ws.merge_cells(f'B{row}:C{row}')
    for i in range(1, 9):
        col = get_column_letter(i)
        if col != 'C':  # Skip merged cell
            ws[f'{col}{row}'].border = thin_border
            if i >= 7:  # Unit Cost and Line Total columns
                ws[f'{col}{row}'] = 0.00
                ws[f'{col}{row}'].number_format = '0.00'
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
        else:
            ws[f'{col}{row}'].border = thin_border

# Payment details section
payment_row = 25
ws['A25'] = 'Payment details:'
ws['A25'].font = header_font

payment_labels = ['BSB:', 'Account:', 'Name:', 'Terms:', 'Email for remittance:']
payment_values = ['325185', '3888190', 'Donna Hattam', 'e.g. 14 days', 'donnacapurso@gmail.com']

for i, (label, value) in enumerate(zip(payment_labels, payment_values)):
    row = 26 + i
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = Font(italic=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'] = value

# Totals section (right side)
totals_data = [
    ('Sub-total', 57.60, light_orange_fill),
    ('Freight', 0.00, light_orange_fill),
    ('Total', 57.60, orange_fill),
    ('GST amount', 0.00, light_orange_fill),
]

for i, (label, value, fill) in enumerate(totals_data):
    row = 25 + i
    ws[f'G{row}'] = label
    ws[f'G{row}'].fill = fill
    ws[f'G{row}'].border = thin_border
    ws[f'G{row}'].alignment = Alignment(horizontal='right')
    ws[f'H{row}'] = value
    ws[f'H{row}'].fill = fill
    ws[f'H{row}'].border = thin_border
    ws[f'H{row}'].number_format = '0.00'
    ws[f'H{row}'].alignment = Alignment(horizontal='right')

# Save the workbook
wb.save('tax_invoice.xlsx')
print("Tax invoice spreadsheet created successfully: tax_invoice.xlsx")
print("ABN field is formatted as text to prevent date conversion issues.")
